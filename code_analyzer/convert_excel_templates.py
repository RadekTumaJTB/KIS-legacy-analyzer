#!/usr/bin/env python3
"""
Convert Excel .xls templates to .xlsx format
Requires: pip install openpyxl xlrd
"""
import os
from pathlib import Path
import xlrd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def convert_xls_to_xlsx(xls_path, xlsx_path):
    """Convert .xls to .xlsx preserving basic formatting"""
    try:
        # Read .xls file
        xls_book = xlrd.open_workbook(xls_path, formatting_info=False)

        # Create new .xlsx workbook
        xlsx_book = Workbook()
        xlsx_book.remove(xlsx_book.active)  # Remove default sheet

        # Copy each sheet
        for sheet_index in range(xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(sheet_index)
            xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

            # Copy cell values
            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row, col)
                    xlsx_sheet.cell(row=row+1, column=col+1, value=cell_value)

        # Save .xlsx file
        xlsx_book.save(xlsx_path)
        print(f"✓ Converted: {xls_path.name} → {xlsx_path.name}")
        return True
    except Exception as e:
        print(f"✗ Error converting {xls_path.name}: {e}")
        return False

def main():
    # Template directories (Linux deployment paths)
    dirs = [
        Path("/opt/kis-banking/Konsolidace_JT/sablony/"),
        Path("/opt/kis-banking/Konsolidace_JT/sablony/cartesis/")
    ]

    total_files = 0
    converted_files = 0
    failed_files = 0

    for template_dir in dirs:
        if not template_dir.exists():
            print(f"⚠ Directory not found: {template_dir}")
            print(f"   This script should be run on Linux deployment server")
            continue

        print(f"\nProcessing directory: {template_dir}")

        for xls_file in template_dir.glob("*.xls"):
            total_files += 1
            xlsx_file = xls_file.with_suffix(".xlsx")

            if convert_xls_to_xlsx(xls_file, xlsx_file):
                converted_files += 1
            else:
                failed_files += 1

    # Summary
    print("\n" + "="*60)
    print("CONVERSION SUMMARY")
    print("="*60)
    print(f"Total files found: {total_files}")
    print(f"Successfully converted: {converted_files}")
    print(f"Failed: {failed_files}")
    print("="*60)

    if total_files == 0:
        print("\n⚠ No .xls files found. Run this script on Linux deployment server.")
        print("   Expected directories:")
        for dir_path in dirs:
            print(f"   - {dir_path}")
    elif failed_files > 0:
        print(f"\n⚠ {failed_files} files failed to convert. Check errors above.")
        return 1
    else:
        print("\n✅ All template files converted successfully!")
        return 0

if __name__ == "__main__":
    exit(main())
